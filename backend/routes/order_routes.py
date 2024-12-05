from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_
from datetime import datetime, timezone
from database import get_db
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from io import BytesIO
import urllib.parse

from models import Order, Event, Payments, OrderItems, AlterationDetails, Affiliation, Author, Product, Form, FormCategory
from models.order import OrderStatus
from schemas.category_schema import AttributeResponse, CategoryResponse
from schemas.order_schema import OrderListResponse, OrderDetailResponse, OrderCreate, OrderStatusUpdate, PaymentInfo, OrderFilterResponse, OrderItemResponse, ProductResponse
from schemas.form_schema import FormResponse, FormRepairResponse
from schemas.alteration_details_schema import AlterationDetailsInfo

router = APIRouter()

# 1. 주문서 리스트 조회 API
@router.get("/orders", response_model=OrderListResponse, summary="주문서 조회(필터)", tags=["주문서 API"])
async def get_orders(
    event_name: Optional[str] = None,
    order_date_from: Optional[datetime] = None,
    order_date_to: Optional[datetime] = None,
    sort: Optional[str] = "order_date_asc",
    search: Optional[str] = None,
    status: Optional[str] = None,
    is_temp: Optional[bool] = None,
    limit: int = 10,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    """
    주문서 조회 API 필터 설명\n
    1.event_name:\n
    설명: 특정 이벤트 이름으로 주문서를 필터링.\n
    사용법: /orders?event_name=Fashion Week\n
    2.status:\n
    설명: 주문 상태로 필터링.\n
        Order_Completed = 'Order Completed' / 주문완료\n
        Packaging_Completed = 'Packaging Completed' / 포장완료\n
        Repair_Received = 'Repair Received' / 수선 접수\n
        Repair_Completed = 'Repair Completed' / 수선 완료\n
        In_delivery = 'In delivery' / 배송중\n
        Delivery_completed = 'Delivery completed' / 배송완료\n
        Receipt_completed = 'Receipt completed' / 수령완료\n
        Accommodation = 'Accommodation' / 숙소\n
        Counsel = 'Counsel' / 상담\n
    사용법: /orders?status=Order_Completed\n
    3.order_date_from 및 order_date_to:\n
    설명: 지정된 기간 내의 주문서만 조회.\n
    사용법: /orders?order_date_from=2023-01-01&order_date_to=2023-12-31\n
    4.sort:\n
    설명: 주문서를 정렬 기준으로 정렬.\n
    옵션: order_date_asc (날짜 오름차순), order_date_desc (날짜 내림차순)\n
    사용법: /orders?sort=order_date_asc\n
    5.search:\n
    설명: [고객 이름, 작성자, 주소, 소속, 연락처, 노트, 수선 노트]로 검색.\n
    사용법: /orders?search=John \n
    6.is_temp:\n
    설명: 일반/임시 주문서 선택 조회\n
    사용법: /orders?is_temp=False\n
    7.limit 및 offset:\n
    설명: 페이징을 위한 필터, limit은 반환할 주문서 수, offset은 시작 위치.\n
    사용법: /orders?limit=10&offset=20\n
    """
    query = db.query(Order).options(
        joinedload(Order.author),
        joinedload(Order.affiliation),
        joinedload(Order.payments),
        joinedload(Order.event).joinedload(Event.form).joinedload(Form.form_repairs),
        joinedload(Order.order_items).joinedload(OrderItems.product).joinedload(Product.attributes),
        joinedload(Order.alteration_details).joinedload(AlterationDetails.form_repair)
    )

    if event_name:
        query = query.join(Event).filter(Event.name == event_name)
    if order_date_from and order_date_to:
        query = query.filter(and_(Order.created_at >= order_date_from, Order.created_at <= order_date_to))
    if search:
        query = query.filter(
            or_(
                Order.groomName.ilike(f"%{search}%"),
                Order.brideName.ilike(f"%{search}%"),
                Order.address.ilike(f"%{search}%"),
                Order.contact.ilike(f"%{search}%"),
                Order.notes.ilike(f"%{search}%"),
                Order.alter_notes.ilike(f"%{search}%"),
                Author.name.ilike(f"%{search}%"),
                Affiliation.name.ilike(f"%{search}%")
            )
        )
    if status:
        query = query.filter(Order.status == status)
    if is_temp is not None:
        query = query.filter(Order.isTemporary == is_temp)

    if sort == "order_date_asc":
        query = query.order_by(Order.created_at.asc())
    elif sort == "order_date_desc":
        query = query.order_by(Order.created_at.desc())

    total_orders = query.count()
    query = query.offset(offset).limit(limit)
    orders = query.all()

    order_list = [
        OrderFilterResponse(
            id=order.id,
            event_id=order.event_id,
            author_id=order.author_id,
            modifier_id=order.modifier_id,
            affiliation_id=order.affiliation_id,
            orderNumber=order.orderNumber,
            event_name=order.event.name,
            form_name=order.event.form.name,
            groomName=order.groomName,
            brideName=order.brideName,
            orderItems=[
                OrderItemResponse(
                    product=ProductResponse(
                        id=item.product_id,
                        name=item.product.name,
                        price=item.product.price
                    ),
                    price=item.price,
                    quantity=item.quantity,
                    attributes=[
                        AttributeResponse(id=item.attribute.id, value=item.attribute.value)
                    ] if item.attribute else []
                ) for item in order.order_items
            ],
            payments=[
                PaymentInfo(
                    payer=payment.payer,
                    payment_date=payment.payment_date,
                    cashAmount=payment.cashAmount,
                    cashCurrency=payment.cashCurrency,
                    cashConversion=payment.cashConversion,
                    cardAmount=payment.cardAmount,
                    cardCurrency=payment.cardCurrency,
                    cardConversion=payment.cardConversion,
                    tradeInAmount=payment.tradeInAmount,
                    tradeInCurrency=payment.tradeInCurrency,
                    tradeInConversion=payment.tradeInConversion,
                    paymentMethod=payment.paymentMethod,
                    notes=payment.notes
                ) for payment in order.payments
            ],
            created_at=order.created_at,
            updated_at=order.updated_at,
            status=order.status,
            contact=order.contact,
            address=order.address,
            collectionMethod=order.collectionMethod,
            notes=order.notes,
            alter_notes=order.alter_notes,
            totalPrice=order.totalPrice,
            advancePayment=order.advancePayment,
            balancePayment=order.balancePayment
        ) for order in orders
    ]

    return OrderListResponse(orders=order_list, total=total_orders)


# 2. 단일 주문서 상세 조회 API
@router.get("/order/{orderID}", response_model=OrderDetailResponse, summary="주문서 상세 조회", tags=["주문서 API"])
async def get_order_detail(orderID: int, db: Session = Depends(get_db)):
    order = db.query(Order).options(
        joinedload(Order.event).joinedload(Event.form).joinedload(Form.form_repairs),
        joinedload(Order.event).joinedload(Event.form).joinedload(Form.form_categories).joinedload(FormCategory.category),
        joinedload(Order.order_items).joinedload(OrderItems.product),
        joinedload(Order.order_items).joinedload(OrderItems.attribute),
        joinedload(Order.payments),
        joinedload(Order.alteration_details).joinedload(AlterationDetails.form_repair)
    ).filter(Order.id == orderID).first()

    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order_detail = OrderDetailResponse(
        id=order.id,
        event_id=order.event_id,
        author_id=order.author_id,
        modifier_id=order.modifier_id,
        affiliation_id=order.affiliation_id,
        orderNumber=order.orderNumber,
        event_name=order.event.name,
        groomName=order.groomName,
        brideName=order.brideName,
        form=FormResponse(
            id=order.event.form.id,
            name=order.event.form.name,
            repairs=[
                FormRepairResponse(
                    id=repair.id,
                    information=repair.information,
                    unit=repair.unit,
                    isAlterable=repair.isAlterable,
                    standards=repair.standards,
                    indexNumber=repair.indexNumber
                ) for repair in sorted(order.event.form.form_repairs, key=lambda r: r.indexNumber)
            ],
            categories=[
                CategoryResponse(id=form_category.category.id, name=form_category.category.name)
                for form_category in order.event.form.form_categories
            ],
            created_at=order.event.form.created_at
        ),
        orderItems=[
            OrderItemResponse(
                product=ProductResponse(
                    id=item.product_id,
                    name=item.product.name,
                    price=item.product.price
                ),
                price=item.price,
                quantity=item.quantity,
                attributes=[
                    AttributeResponse(id=item.attribute.id, value=item.attribute.value)
                ] if item.attribute else []
            ) for item in order.order_items
        ],
        payments=[
            PaymentInfo(
                payer=payment.payer,
                payment_date=payment.payment_date,
                cashAmount=payment.cashAmount,
                cashCurrency=payment.cashCurrency,
                cashConversion=payment.cashConversion,
                cardAmount=payment.cardAmount,
                cardCurrency=payment.cardCurrency,
                cardConversion=payment.cardConversion,
                tradeInAmount=payment.tradeInAmount,
                tradeInCurrency=payment.tradeInCurrency,
                tradeInConversion=payment.tradeInConversion,
                paymentMethod=payment.paymentMethod,
                notes=payment.notes
            ) for payment in order.payments
        ],
        alteration_details=[
            AlterationDetailsInfo(
                form_repair_id=detail.form_repair_id,
                figure=detail.figure,
                alterationFigure=detail.alterationFigure,
            ) for detail in order.alteration_details
        ],
        created_at=order.created_at,
        updated_at=order.updated_at,
        status=order.status,
        contact=order.contact,
        address=order.address,
        collectionMethod=order.collectionMethod,
        notes=order.notes,
        alter_notes=order.alter_notes,
        totalPrice=order.totalPrice,
        advancePayment=order.advancePayment,
        balancePayment=order.balancePayment
    )

    return order_detail


# 주문 상태 업데이트 API
@router.put("/orders/{orderID}/{order_status}", response_model=OrderStatusUpdate, summary="주문 상태 업데이트", tags=["주문서 API"])
async def update_order_status(
    orderID: int,
    order_status: str,
    db: Session = Depends(get_db)
):
    """
        Order_Completed = 'Order Completed' / 주문완료\n
        Packaging_Completed = 'Packaging Completed' / 포장완료\n
        Repair_Received = 'Repair Received' / 수선 접수\n
        Repair_Completed = 'Repair Completed' / 수선 완료\n
        In_delivery = 'In delivery' / 배송중\n
        Delivery_completed = 'Delivery completed' / 배송완료\n
        Receipt_completed = 'Receipt completed' / 수령완료\n
        Accommodation = 'Accommodation' / 숙소\n
        Counsel = 'Counsel' / 상담\n
    """
    # 주문서 조회
    order = db.query(Order).filter(Order.id == orderID).first()

    if not order:
        raise HTTPException(status_code=404, detail="주문서를 찾을 수 없습니다.")

    # 주문 상태 유효성 검사
    try:
        new_status = OrderStatus(order_status)
    except ValueError:
        raise HTTPException(status_code=400, detail="잘못된 주문 상태입니다.")

    # 주문 상태 업데이트
    order.status = new_status
    order.updated_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(order)

    return {
        "id": order.id,
        "status": order.status,
        "updated_at": order.updated_at
    }

@router.post("/order/save", summary="주문서 생성", status_code=status.HTTP_201_CREATED, tags=["주문서 API"])
async def create_order(order: OrderCreate, db: Session = Depends(get_db), is_temp: bool = False):
    """
    새로운 주문서 생성
    """
    try:
        # 임시 저장이 아닌 경우에만 주문번호 생성
        order_number = None
        if not is_temp:  
            today_prefix = datetime.now(timezone.utc).strftime("%y%m%d")
            latest_order = (
                db.query(Order)
                .filter(Order.orderNumber.like(f"{today_prefix}-%"))
                .order_by(Order.orderNumber.desc())
                .first()
            )
            if latest_order:
                last_sequence = int(latest_order.orderNumber.split("-")[1])
                next_sequence = f"{last_sequence + 1:03d}"
            else:
                next_sequence = "001"
            order_number = f"{today_prefix}-{next_sequence}"

        # 새로운 주문서 생성
        new_order = Order(
            event_id=order.event_id,
            author_id=order.author_id,
            modifier_id=order.modifier_id,
            orderNumber=order_number,  # 생성된 주문번호 할당
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
            status=order.status,
            groomName=order.groomName,
            brideName=order.brideName,
            contact=order.contact,
            affiliation_id=order.affiliation_id,
            address=order.address,
            collectionMethod=order.collectionMethod,
            notes=order.notes,
            alter_notes=order.alter_notes,
            totalPrice=order.totalPrice,
            advancePayment=order.advancePayment,
            balancePayment=order.balancePayment,
            isTemporary=is_temp
        )

        db.add(new_order)
        db.flush()  # 데이터베이스에 추가하고 ID 확보

        # 주문 상품 정보 (OrderItems) 저장
        for order_item in order.orderItems:
            new_item = OrderItems(
                order_id=new_order.id,
                product_id=order_item.product_id,
                attribute_id=order_item.attributes_id,
                quantity=order_item.quantity,
                price=order_item.price
            )
            db.add(new_item)
        
        # 결제 정보 저장
        if order.payments:
            for payment in order.payments:
                new_payment = Payments(
                    order_id=new_order.id,
                    payer=payment.payer,
                    payment_date=payment.payment_date,
                    cashAmount=payment.cashAmount,
                    cashCurrency=payment.cashCurrency,
                    cashConversion=payment.cashConversion,
                    cardAmount=payment.cardAmount,
                    cardCurrency=payment.cardCurrency,
                    cardConversion=payment.cardConversion,
                    tradeInAmount=payment.tradeInAmount,
                    tradeInCurrency=payment.tradeInCurrency,
                    tradeInConversion=payment.tradeInConversion,
                    paymentMethod=payment.paymentMethod,
                    notes=payment.notes
                )
                db.add(new_payment)

        # 수선 정보 저장
        if order.alteration_details:
            for alteration in order.alteration_details:
                new_alteration = AlterationDetails(
                    order_id=new_order.id,
                    form_repair_id=alteration.form_repair_id,
                    figure=alteration.figure,
                    alterationFigure=alteration.alterationFigure,
                )
                db.add(new_alteration)

        # 모든 데이터 커밋
        db.commit()
        return {
            "message": "Order saved successfully!",
            "order_id": new_order.id,
            "orderNumber": new_order.orderNumber if not is_temp else None 
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"주문 생성 중 오류가 발생했습니다: {str(e)}")

@router.put("/order/save/{order_id}", summary="주문서 수정", tags=["주문서 API"])
async def update_order(order_id: int, order: OrderCreate, db: Session = Depends(get_db), is_temp: bool = False):
    """
    기존 주문서 수정
    """
    try:
        existing_order = db.query(Order).filter(Order.id == order_id).first()
        if not existing_order:
            raise HTTPException(status_code=404, detail="Order not found")

        # orderNumber가 None이고 is_temp가 False일 경우 새로운 주문번호 생성
        if not is_temp and not existing_order.orderNumber:
            today_prefix = datetime.now(timezone.utc).strftime("%y%m%d")
            latest_order = (
                db.query(Order)
                .filter(Order.orderNumber.like(f"{today_prefix}-%"))
                .order_by(Order.orderNumber.desc())
                .first()
            )
            if latest_order:
                last_sequence = int(latest_order.orderNumber.split("-")[1])
                next_sequence = f"{last_sequence + 1:03d}"
            else:
                next_sequence = "001"
            existing_order.orderNumber = f"{today_prefix}-{next_sequence}"

        # 주문 정보 수정
        existing_order.event_id = order.event_id
        existing_order.author_id = order.author_id
        existing_order.modifier_id = order.modifier_id
        existing_order.updated_at = datetime.now(timezone.utc)
        existing_order.status = order.status
        existing_order.groomName = order.groomName
        existing_order.brideName = order.brideName
        existing_order.contact = order.contact
        existing_order.affiliation_id = order.affiliation_id
        existing_order.address = order.address
        existing_order.collectionMethod = order.collectionMethod
        existing_order.notes = order.notes
        existing_order.alter_notes = order.alter_notes
        existing_order.totalPrice = order.totalPrice
        existing_order.advancePayment = order.advancePayment
        existing_order.balancePayment = order.balancePayment
        existing_order.isTemporary = is_temp

        db.flush()

        # 상품 정보 (OrderItems) 업데이트 (삭제 후 재생성)
        db.query(OrderItems).filter(OrderItems.order_id == existing_order.id).delete()
        for order_item in order.orderItems:
            new_item = OrderItems(
                order_id=existing_order.id,
                product_id=order_item.product_id,
                attribute_id=order_item.attributes_id,
                quantity=order_item.quantity,
                price=order_item.price
            )
            db.add(new_item)

        # 결제 정보 업데이트
        for payment in order.payments:
            existing_payment = db.query(Payments).filter(
                and_(Payments.order_id == existing_order.id, Payments.paymentMethod == payment.paymentMethod)
            ).first()
            if existing_payment:
                existing_payment.payer = payment.payer
                existing_payment.payment_date = payment.payment_date
                existing_payment.cashAmount = payment.cashAmount
                existing_payment.cashCurrency = payment.cashCurrency
                existing_payment.cashConversion = payment.cashConversion
                existing_payment.cardAmount = payment.cardAmount
                existing_payment.cardCurrency = payment.cardCurrency
                existing_payment.cardConversion = payment.cardConversion
                existing_payment.tradeInAmount = payment.tradeInAmount
                existing_payment.tradeInCurrency = payment.tradeInCurrency
                existing_payment.tradeInConversion = payment.tradeInConversion
                existing_payment.notes = payment.notes
            else:
                new_payment = Payments(
                    order_id=existing_order.id,
                    payer=payment.payer,
                    payment_date=payment.payment_date,
                    cashAmount=payment.cashAmount,
                    cashCurrency=payment.cashCurrency,
                    cashConversion=payment.cashConversion,
                    cardAmount=payment.cardAmount,
                    cardCurrency=payment.cardCurrency,
                    cardConversion=payment.cardConversion,
                    tradeInAmount=payment.tradeInAmount,
                    tradeInCurrency=payment.tradeInCurrency,
                    tradeInConversion=payment.tradeInConversion,
                    paymentMethod=payment.paymentMethod,
                    notes=payment.notes
                )
                db.add(new_payment)

        # 수선 정보 업데이트
        for alteration in order.alteration_details:
            existing_alteration = db.query(AlterationDetails).filter(
                and_(AlterationDetails.order_id == existing_order.id, AlterationDetails.form_repair_id == alteration.form_repair_id)
            ).first()
            if existing_alteration:
                existing_alteration.figure = alteration.figure
                existing_alteration.alterationFigure = alteration.alterationFigure
            else:
                new_alteration = AlterationDetails(
                    order_id=existing_order.id,
                    form_repair_id=alteration.form_repair_id,
                    figure=alteration.figure,
                    alterationFigure=alteration.alterationFigure,
                )
                db.add(new_alteration)
        # 모든 데이터 커밋
        db.commit()
        return {"message": "Order updated successfully!", "order_id": existing_order.id}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"주문 수정 중 오류가 발생했습니다: {str(e)}")

# 임시 저장 주문서 API
@router.post("/temp/order/save", summary="임시 주문서 생성", status_code=status.HTTP_201_CREATED, tags=["주문서 API"])
async def create_temp_order(order: OrderCreate, db: Session = Depends(get_db)):
    return await create_order(order, db, is_temp=True)

@router.put("/temp/order/save/{order_id}", summary="임시 주문서 수정", tags=["주문서 API"])
async def update_temp_order(order_id: int, order: OrderCreate, db: Session = Depends(get_db)):
    return await update_order(order_id, order, db, is_temp=True)

@router.delete("/order/{order_id}", summary="주문서 삭제", status_code=status.HTTP_200_OK, tags=["주문서 API"])
async def delete_order(order_id: int, db: Session = Depends(get_db)):
    """
    주문서 삭제 API
    - 해당 주문서와 관련된 OrderItems, Payments, AlterationDetails 테이블의 데이터도 함께 삭제합니다.
    """
    # 주문서 검색
    existing_order = db.query(Order).filter(Order.id == order_id).first()

    if not existing_order:
        raise HTTPException(status_code=404, detail="주문서를 찾을 수 없습니다.")
    # OrderItems 삭제
    db.query(OrderItems).filter(OrderItems.order_id == order_id).delete()
    # Payments 삭제
    db.query(Payments).filter(Payments.order_id == order_id).delete()
    # AlterationDetails 삭제
    db.query(AlterationDetails).filter(AlterationDetails.order_id == order_id).delete()
    # 주문서 삭제
    db.delete(existing_order)
    # 변경 사항 커밋
    db.commit()
    return {"message": "주문서 및 관련 데이터가 성공적으로 삭제되었습니다.", "order_id": order_id}

@router.get("/orders/download", summary="Excel 다운로드", tags=["주문서 API"])
async def download_orders(
    event_name: Optional[str] = None,
    order_date_from: Optional[datetime] = None,
    order_date_to: Optional[datetime] = None,
    sort: Optional[str] = "order_date_asc",
    search: Optional[str] = None,
    status: Optional[str] = None,
    is_temp: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    """
    주문서 Excel 다운로드 API 설명
    이 API는 주문서 조회 필터를 적용하여 주문서 목록을 Excel 파일로 다운로드할 수 있게 합니다.

    필터 설명:
    1. event_name:
       - 설명: 특정 이벤트 이름으로 주문서를 필터링합니다.
       - 사용법: /orders/download?event_name=Fashion Week

    2. status:
       - 설명: 주문 상태로 필터링합니다.
       - 사용 가능한 상태:
            - Order_Completed = 'Order Completed' / 주문완료
            - Packaging_Completed = 'Packaging Completed' / 포장완료
            - Repair_Received = 'Repair Received' / 수선 접수
            - Repair_Completed = 'Repair Completed' / 수선 완료
            - In_delivery = 'In delivery' / 배송중
            - Delivery_completed = 'Delivery completed' / 배송완료
            - Receipt_completed = 'Receipt completed' / 수령완료
            - Accommodation = 'Accommodation' / 숙소
       - 사용법: /orders/download?status=Order_Completed

    3. order_date_from 및 order_date_to:
       - 설명: 지정된 기간 내의 주문서만 조회합니다.
       - 사용법: /orders/download?order_date_from=2023-01-01&order_date_to=2023-12-31

    4. sort:
       - 설명: 주문서를 정렬 기준으로 정렬합니다.
       - 옵션: order_date_asc (날짜 오름차순), order_date_desc (날짜 내림차순)
       - 사용법: /orders/download?sort=order_date_asc

    5. search:
       - 설명: 고객 이름, 작성자, 주소, 소속으로 검색할 수 있습니다.
       - 사용법: /orders/download?search=John 

    6. is_temp:
       - 설명: 일반/임시 주문서를 선택하여 조회합니다.
       - 사용법: /orders/download?is_temp=False

    엑셀 파일 다운로드:
    - API 호출 후 응답으로 생성된 엑셀 파일이 다운로드됩니다.
    - 파일명은 'orders.xlsx'로 제공됩니다.
    """
    # 주문서 조회와 동일한 쿼리 로직 재사용
    query = db.query(Order).options(
        joinedload(Order.author),
        joinedload(Order.modifier),  # 수정자 로드 추가
        joinedload(Order.affiliation),
        joinedload(Order.payments),
        joinedload(Order.event).joinedload(Event.form),
        joinedload(Order.order_items).joinedload(OrderItems.product).joinedload(Product.attributes)
    )

    # 필터 적용
    if event_name:
        query = query.join(Event).filter(Event.name == event_name)

    if order_date_from and order_date_to:
        query = query.filter(and_(Order.created_at >= order_date_from, Order.created_at <= order_date_to))

    if search:
        query = query.filter(
            or_(
                Order.orderName.ilike(f"%{search}%"),
                Order.address.ilike(f"%{search}%"),
                Order.author.name.ilike(f"%{search}%"),
                Order.affiliation.name.ilike(f"%{search}%")
            )
        )

    if status:
        query = query.filter(Order.status == status)

    if is_temp is not None:
        query = query.filter(Order.isTemporary == is_temp)

    # 정렬 옵션 처리
    if sort == "order_date_asc":
        query = query.order_by(Order.created_at.asc())
    elif sort == "order_date_desc":
        query = query.order_by(Order.created_at.desc())

    # 모든 필터링된 주문서 가져오기
    orders = query.all()

    # 데이터프레임 생성
    rows = []
    for order in orders:
        for payment in order.payments:
            rows.append({
                "Event Name": order.event.name if order.event else None,
                "Author Name": order.author.name if order.author else None,
                "Modifier Name": order.modifier.name if order.modifier else None,  # 수정자 이름 추가
                "Order Name": order.orderName,
                "Contact": order.contact,
                "Affiliation Name": order.affiliation.name if order.affiliation else None,
                "Collection Method": order.collectionMethod,
                "Status": order.status,
                "Created At": order.created_at,
                "Updated At": order.updated_at,  # 수정 시간 추가
                "Total Price": order.totalPrice,
                "Total Payment": order.advancePayment + order.balancePayment,
                "Payment Date": payment.payment_date,
                "Payment Method": payment.paymentMethod,
                "Address": order.address,
                "Notes": payment.notes,
            })

    df = pd.DataFrame(rows)

    # Excel 파일 생성 및 스타일링
    wb = Workbook()
    ws = wb.active

    # 칼럼 헤더 설정 (2번째 줄)
    headers = [
        "행사명", "작성자", "수정자", "주문자", "연락처",  # 수정자 컬럼 추가
        "소속", "수령 방법", "주문 상태", 
        "주문서 생성날짜", "주문서 수정날짜",  # 수정 날짜 추가
        "총 주문 금액", "총 결제 금액", 
        "결제 날짜", "결제 방식", "주소", "비고"
    ]

    for c_idx, header in enumerate(headers, 1):
        ws.cell(row=2, column=c_idx, value=header)

    # 데이터 입력 (3번째 줄부터)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 헤더 행에 필터 추가 (2번째 줄부터)
    ws.auto_filter.ref = f"A2:P{len(rows) + 2}"  # 필터 범위를 2번째 줄부터 설정

    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                print(f"Error calculating length for cell {cell}: {e}")

        # 열 문자 가져오기 (첫 번째 셀을 사용)
        if max_length > 0:  # max_length가 0보다 클 경우에만 너비 설정
            column_letter = column[0].column_letter  # 열 문자 가져오기
            adjusted_width = (max_length + 2)  # 여유 공간 추가
            ws.column_dimensions[column_letter].width = adjusted_width

    # 행사명 설정 (A1) 및 병합
    ws.merge_cells('A1:P1')  # 행사명 셀 병합
    ws['A1'] = f"행사명: {event_name or '전체'} 주문서 목록"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    # 엑셀 파일 버퍼에 저장
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # 현재 시간을 포맷하여 파일 이름에 추가
    current_time = datetime.now().strftime("%Y%m%d")  # 예: 20230924_153045
    file_name = f"{event_name or '전체 주문서'}_{current_time}.xlsx"
    file_name = urllib.parse.quote(file_name)  # URL 인코딩

    # 다운로드 응답 반환
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{file_name}"}
    )
